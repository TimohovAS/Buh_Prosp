from __future__ import annotations

import base64
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

from backend.models import Enterprise, Project, WorkDiaryEntry

MONEY_STEP = Decimal("0.01")
ZERO = Decimal("0")
DEFAULT_MATERIAL_MULTIPLIER = Decimal("1.2")
ASSET_LOGO_PATH = Path(__file__).with_name("assets") / "prospel_logo.jpg"

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
LIGHT_GRAY = "E7E6E6"
TEXT = "1F2937"
WHITE = "FFFFFF"
RED = "C00000"

THIN_GRAY = Side(style="thin", color="9CA3AF")
MEDIUM_NAVY = Side(style="medium", color=NAVY)


@dataclass(frozen=True)
class ProposalLine:
    work_date: date
    description: str
    unit: str
    quantity: Decimal
    unit_price: Decimal
    amount: Decimal
    line_type: str


def _dec(value) -> Decimal:
    if value in (None, ""):
        return ZERO
    return Decimal(str(value))


def _money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_STEP, rounding=ROUND_HALF_UP)


def _clean(value: str | None) -> str:
    return str(value or "").strip()


def _client_language(project: Project) -> str:
    language = _clean(project.client.document_language if project.client else "").lower()
    return language if language in {"sr", "ru"} else "sr"


LABELS = {
    "sr": {
        "proposal": "PREDLOG ZA FAKTURISANJE",
        "specification": "SPECIFIKACIJA RADOVA I MATERIJALA",
        "customer": "NARUČILAC",
        "document": "PODACI O PREDLOGU",
        "document_date": "Datum dokumenta:",
        "work_period": "Period radova:",
        "project": "Projekat:",
        "currency": "Valuta:",
        "ordinal": "Red. br.",
        "date": "Datum",
        "item": "Naziv usluge / materijala",
        "unit": "Jed. mere",
        "quantity": "Kol.",
        "price": "Cena",
        "amount": "Ukupno",
        "work": "Radovi",
        "material": "Materijal",
        "adjustment": "Korekcija dogovorene cene",
        "piece": "stavka",
        "total": "UKUPNO PREDLOŽENO ZA FAKTURISANJE:",
        "payment": "Rok plaćanja: prema ugovoru ili dogovoru sa naručiocem.",
        "disclaimer": (
            "Ovaj dokument je predlog za fakturisanje i specifikacija izvedenih radova i "
            "utrošenog materijala. Nije faktura niti zahtev za plaćanje. PDV i poreska "
            "kategorija biće iskazani u konačnoj elektronskoj fakturi."
        ),
        "prepared_by": "Sastavio:",
        "pib": "PIB",
        "mb": "MB",
        "bank_account": "ŽR",
    },
    "ru": {
        "proposal": "ПРЕДЛОЖЕНИЕ К ФАКТУРИРОВАНИЮ",
        "specification": "СПЕЦИФИКАЦИЯ РАБОТ И МАТЕРИАЛОВ",
        "customer": "ЗАКАЗЧИК",
        "document": "ДАННЫЕ ПРЕДЛОЖЕНИЯ",
        "document_date": "Дата документа:",
        "work_period": "Период работ:",
        "project": "Проект:",
        "currency": "Валюта:",
        "ordinal": "№",
        "date": "Дата",
        "item": "Наименование работы / материала",
        "unit": "Ед.",
        "quantity": "Кол.",
        "price": "Цена",
        "amount": "Сумма",
        "work": "Работы",
        "material": "Материал",
        "adjustment": "Корректировка договорной цены",
        "piece": "позиция",
        "total": "ИТОГО ПРЕДЛОЖЕНО К ФАКТУРИРОВАНИЮ:",
        "payment": "Срок оплаты: согласно договору или договоренности с заказчиком.",
        "disclaimer": (
            "Этот документ является предложением к фактурированию и спецификацией выполненных "
            "работ и использованных материалов. Он не является фактурой или требованием оплаты. "
            "НДС и налоговая категория будут указаны в окончательной электронной фактуре."
        ),
        "prepared_by": "Составил:",
        "pib": "PIB",
        "mb": "MB",
        "bank_account": "Счет",
    },
}


def _proposal_lines(entries: list[WorkDiaryEntry], labels: dict[str, str]) -> list[ProposalLine]:
    lines: list[ProposalLine] = []
    for entry in entries:
        duration = _dec(entry.duration_hours)
        labor_rate = _dec(entry.team_billing_hourly_rate_snapshot)
        labor_amount = _money(duration * labor_rate)
        lines.append(
            ProposalLine(
                work_date=entry.date,
                description=f"{labels['work']}: {_clean(entry.description)}",
                unit="h",
                quantity=duration,
                unit_price=labor_rate,
                amount=labor_amount,
                line_type="work",
            )
        )

        multiplier = _dec(entry.material_billing_multiplier) or DEFAULT_MATERIAL_MULTIPLIER
        material_amount = ZERO
        for material in entry.materials:
            billed_amount = _money(_dec(material.amount) * multiplier)
            material_amount += billed_amount
            quantity = _dec(material.quantity)
            if quantity <= 0:
                quantity = Decimal("1")
            unit_price = _money(billed_amount / quantity) if quantity else billed_amount
            lines.append(
                ProposalLine(
                    work_date=entry.date,
                    description=f"{labels['material']}: {_clean(material.description)}",
                    unit=_clean(material.unit) or labels["piece"],
                    quantity=quantity,
                    unit_price=unit_price,
                    amount=billed_amount,
                    line_type="material",
                )
            )

        calculated_amount = labor_amount + material_amount
        if entry.billable_amount_override is not None:
            adjustment = _money(_dec(entry.billable_amount_override) - calculated_amount)
            if adjustment != ZERO:
                lines.append(
                    ProposalLine(
                        work_date=entry.date,
                        description=labels["adjustment"],
                        unit=labels["piece"],
                        quantity=Decimal("1"),
                        unit_price=adjustment,
                        amount=adjustment,
                        line_type="adjustment",
                    )
                )
    return lines


def _decode_emblem(data_url: str | None) -> BytesIO | None:
    value = _clean(data_url)
    if not value or "," not in value:
        return None
    header, encoded = value.split(",", 1)
    if ";base64" not in header:
        return None
    try:
        return BytesIO(base64.b64decode(encoded))
    except (ValueError, TypeError):
        return None


def _add_logo(sheet, enterprise: Enterprise) -> None:
    source = _decode_emblem(enterprise.emblem_data_url)
    if source is None and ASSET_LOGO_PATH.exists():
        source = ASSET_LOGO_PATH
    if source is None:
        return
    try:
        logo = ExcelImage(source)
    except (OSError, ValueError):
        return
    logo.width = 86
    logo.height = 86
    sheet.add_image(logo, "A1")


def _set_merged_value(
    sheet,
    cell_range: str,
    value,
    *,
    font: Font | None = None,
    alignment: Alignment | None = None,
    fill: PatternFill | None = None,
    border: Border | None = None,
) -> None:
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = value
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border


def _party_identifier(labels: dict[str, str], *, pib: str | None, mb: str | None) -> str:
    parts = []
    if _clean(pib):
        parts.append(f"{labels['pib']}: {_clean(pib)}")
    if _clean(mb):
        parts.append(f"{labels['mb']}: {_clean(mb)}")
    return "   |   ".join(parts)


def build_work_diary_proposal_xlsx(
    *,
    enterprise: Enterprise,
    project: Project,
    entries: list[WorkDiaryEntry],
    prepared_by: str,
    document_date: date,
) -> BytesIO:
    language = _client_language(project)
    labels = LABELS[language]
    client = project.client
    lines = _proposal_lines(entries, labels)
    period_from = min(entry.date for entry in entries)
    period_to = max(entry.date for entry in entries)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Predlog" if language == "sr" else "Предложение"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A17"

    for column, width in {
        "A": 7,
        "B": 13,
        "C": 48,
        "D": 11,
        "E": 16,
        "F": 16,
        "G": 18,
    }.items():
        sheet.column_dimensions[column].width = width

    for row in range(1, 6):
        sheet.row_dimensions[row].height = 19
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 26
    sheet.row_dimensions[13].height = 28
    sheet.row_dimensions[14].height = 28
    sheet.row_dimensions[16].height = 34
    sheet.row_dimensions[10].height = 36

    title_font = Font(name="Arial", size=26, bold=True, color=NAVY)
    subtitle_font = Font(name="Arial", size=12, bold=True, color=BLUE)
    body_font = Font(name="Arial", size=10, color=TEXT)
    small_font = Font(name="Arial", size=9, color=TEXT)
    section_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    table_header_font = Font(name="Arial", size=9, bold=True, color=WHITE)

    _set_merged_value(sheet, "A1:B5", "")
    _add_logo(sheet, enterprise)
    _set_merged_value(
        sheet,
        "C1:E2",
        _clean(enterprise.name) or "ProspEl",
        font=title_font,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _set_merged_value(
        sheet,
        "C3:E3",
        labels["proposal"],
        font=subtitle_font,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _set_merged_value(
        sheet,
        "C4:E5",
        labels["specification"],
        font=Font(name="Arial", size=10, bold=True, color=TEXT),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    issuer_rows = [
        _clean(enterprise.name),
        _clean(enterprise.address),
        _party_identifier(labels, pib=enterprise.pib, mb=enterprise.maticni_broj),
        (f"{labels['bank_account']}: {_clean(enterprise.bank_account)}" if _clean(enterprise.bank_account) else ""),
        _clean(enterprise.bank_name),
    ]
    for row, value in enumerate(issuer_rows, start=1):
        _set_merged_value(
            sheet,
            f"F{row}:G{row}",
            value,
            font=small_font,
            alignment=Alignment(horizontal="right", vertical="center", wrap_text=True),
        )

    section_fill = PatternFill("solid", fgColor=NAVY)
    _set_merged_value(
        sheet,
        "A7:D7",
        labels["customer"],
        font=section_font,
        alignment=Alignment(horizontal="left", vertical="center"),
        fill=section_fill,
    )
    _set_merged_value(
        sheet,
        "E7:G7",
        labels["document"],
        font=section_font,
        alignment=Alignment(horizontal="left", vertical="center"),
        fill=section_fill,
    )

    customer_rows = [
        _clean(client.name if client else ""),
        _clean(client.address if client else ""),
        _party_identifier(
            labels,
            pib=client.pib if client else None,
            mb=client.maticni_broj if client else None,
        ),
        _clean(client.contact if client else ""),
    ]
    for offset, value in enumerate(customer_rows, start=8):
        _set_merged_value(
            sheet,
            f"A{offset}:D{offset}",
            value,
            font=Font(name="Arial", size=10, bold=offset == 8, color=TEXT),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )

    period_text = (
        period_from.strftime("%d.%m.%Y")
        if period_from == period_to
        else f"{period_from.strftime('%d.%m.%Y')} - {period_to.strftime('%d.%m.%Y')}"
    )
    document_rows = [
        (labels["document_date"], document_date.strftime("%d.%m.%Y")),
        (labels["work_period"], period_text),
        (
            labels["project"],
            (f"{_clean(project.name)} ({_clean(project.code)})" if _clean(project.code) else _clean(project.name)),
        ),
        (labels["currency"], "RSD"),
    ]
    for row, (label, value) in enumerate(document_rows, start=8):
        sheet[f"E{row}"] = label
        sheet[f"E{row}"].font = Font(name="Arial", size=9, bold=True, color=TEXT)
        sheet[f"E{row}"].alignment = Alignment(vertical="center")
        _set_merged_value(
            sheet,
            f"F{row}:G{row}",
            value,
            font=small_font,
            alignment=Alignment(horizontal="right", vertical="center", wrap_text=True),
        )

    _set_merged_value(
        sheet,
        "A13:G13",
        labels["specification"],
        font=Font(name="Arial", size=16, bold=True, color=NAVY),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    object_name = _clean(project.work_diary_meta.object_name if project.work_diary_meta else "")
    _set_merged_value(
        sheet,
        "A14:G14",
        object_name or _clean(project.name),
        font=Font(name="Arial", size=10, bold=True, italic=True, color=TEXT),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    headers = [
        labels["ordinal"],
        labels["date"],
        labels["item"],
        labels["unit"],
        labels["quantity"],
        f"{labels['price']} (RSD)",
        f"{labels['amount']} (RSD)",
    ]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=16, column=column, value=value)
        cell.font = table_header_font
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=MEDIUM_NAVY, bottom=MEDIUM_NAVY, left=THIN_GRAY, right=THIN_GRAY)

    first_line_row = 17
    for index, line in enumerate(lines, start=1):
        row = first_line_row + index - 1
        sheet.cell(row=row, column=1, value=index)
        sheet.cell(row=row, column=2, value=line.work_date)
        sheet.cell(row=row, column=3, value=line.description)
        sheet.cell(row=row, column=4, value=line.unit)
        sheet.cell(row=row, column=5, value=float(line.quantity))
        sheet.cell(row=row, column=6, value=float(line.unit_price))
        sheet.cell(row=row, column=7, value=float(line.amount))
        sheet.row_dimensions[row].height = 30 if len(line.description) < 95 else 46

        row_fill = PatternFill("solid", fgColor=LIGHT_BLUE if line.line_type == "material" else WHITE)
        if line.line_type == "adjustment":
            row_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        for column in range(1, 8):
            cell = sheet.cell(row=row, column=column)
            cell.font = body_font
            cell.fill = row_fill
            cell.border = Border(bottom=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY)
            cell.alignment = Alignment(
                horizontal="left" if column == 3 else "center",
                vertical="center",
                wrap_text=column == 3,
            )
        sheet.cell(row=row, column=2).number_format = "dd.mm.yyyy"
        sheet.cell(row=row, column=5).number_format = "General"
        sheet.cell(row=row, column=6).number_format = "#,##0.00;[Red]-#,##0.00"
        sheet.cell(row=row, column=7).number_format = "#,##0.00;[Red]-#,##0.00"
        sheet.cell(row=row, column=7).font = Font(name="Arial", size=10, bold=True, color=TEXT)

    last_line_row = first_line_row + len(lines) - 1
    total_row = last_line_row + 1
    _set_merged_value(
        sheet,
        f"A{total_row}:F{total_row}",
        labels["total"],
        font=Font(name="Arial", size=11, bold=True, color=WHITE),
        alignment=Alignment(horizontal="right", vertical="center"),
        fill=PatternFill("solid", fgColor=NAVY),
        border=Border(top=MEDIUM_NAVY, bottom=MEDIUM_NAVY),
    )
    total_cell = sheet.cell(row=total_row, column=7)
    total_cell.value = f"=SUM(G{first_line_row}:G{last_line_row})"
    total_cell.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    total_cell.fill = PatternFill("solid", fgColor=NAVY)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = Border(top=MEDIUM_NAVY, bottom=MEDIUM_NAVY, right=MEDIUM_NAVY)
    total_cell.number_format = '#,##0.00 "RSD";[Red]-#,##0.00 "RSD"'
    sheet.row_dimensions[total_row].height = 24

    payment_row = total_row + 2
    _set_merged_value(
        sheet,
        f"A{payment_row}:G{payment_row}",
        labels["payment"],
        font=Font(name="Arial", size=10, bold=True, color=TEXT),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        fill=PatternFill("solid", fgColor=LIGHT_GRAY),
        border=Border(top=THIN_GRAY, bottom=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY),
    )
    sheet.row_dimensions[payment_row].height = 23

    disclaimer_row = payment_row + 1
    _set_merged_value(
        sheet,
        f"A{disclaimer_row}:G{disclaimer_row + 1}",
        labels["disclaimer"],
        font=Font(name="Arial", size=9, italic=True, color=TEXT),
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        border=Border(bottom=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY),
    )
    sheet.row_dimensions[disclaimer_row].height = 28
    sheet.row_dimensions[disclaimer_row + 1].height = 18

    signature_row = disclaimer_row + 4
    _set_merged_value(
        sheet,
        f"E{signature_row}:G{signature_row}",
        f"{labels['prepared_by']} {_clean(prepared_by)}",
        font=body_font,
        alignment=Alignment(horizontal="center", vertical="center"),
        border=Border(top=THIN_GRAY),
    )

    sheet.print_title_rows = "1:16"
    sheet.print_area = f"A1:G{signature_row + 1}"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(left=0.3, right=0.3, top=0.35, bottom=0.4, header=0.1, footer=0.1)
    sheet.oddFooter.center.text = "ProspEl"
    sheet.oddFooter.right.text = "Strana &P / &N"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def proposal_filename(project: Project, entries: list[WorkDiaryEntry]) -> str:
    period_from = min(entry.date for entry in entries)
    period_to = max(entry.date for entry in entries)
    raw_project = _clean(project.code) or _clean(project.name) or f"project_{project.id}"
    safe_project = re.sub(r"[^A-Za-z0-9_-]+", "_", raw_project).strip("_") or f"project_{project.id}"
    return f"predlog_fakturisanja_{safe_project}_{period_from.isoformat()}_{period_to.isoformat()}.xlsx"
